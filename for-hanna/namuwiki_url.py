#이 파이썬 파일은 제목과 링크를 연결시켜서 하나의 엑셀 파일로 만들어 준다.
from matplotlib.pyplot import title
from openpyxl.styles import Font
from selenium import webdriver
from openpyxl import load_workbook, Workbook



chromedriver = "C:\manna\chromedriver.exe"
driver = webdriver.Chrome(chromedriver)
url = "https://namu.wiki/w/%EB%84%A4%EC%9D%B4%EB%B2%84%20%EC%9B%B9%ED%88%B0/%EC%97%B0%EC%9E%AC%20%EB%AA%A9%EB%A1%9D"
driver.get(url)

webtoon_data = {}
webtoon_title_list = []
xlsx_name = "webtoon_mon"

webtoon_detail_list_xpath_base = "//*[@id='wd8ZpUELU']/article/div[4]/div/div/div/div/div/div[6]/div/div/div/div/div/div/div/div/div/div[6]/div[1]/div/div[7]/div[2]/table/tbody/tr/td/div"

webtoon_title = driver.find_elements_by_xpath(webtoon_detail_list_xpath_base + "/a")

for index, kk in enumerate(webtoon_title):
    webtoon_data[kk.text] = kk.get_attribute("href")
    webtoon_title_list.append(kk.text)


#print(webtoon_data)


driver.quit()
    
#엑셀파일 쓰기
work_book = Workbook()

#시트 생성
#work_book.create_sheet(index=0, title= 'naver_fri')


#시트 입력
sheet = work_book.active
sheet.column_dimensions["A"].width = 45

for kk in range(1, len(webtoon_data)):
    sheet['A' + str(kk)] = webtoon_title_list[kk-1]
    sheet['A' + str(kk)].font = Font(name="나눔고딕", color="000000")
    sheet['B' + str(kk)] = webtoon_data[webtoon_title_list[kk-1]]
    sheet['B' + str(kk)].font = Font(name = "나눔고딕", color = "000000")

work_book.save(xlsx_name + ".xlsx")