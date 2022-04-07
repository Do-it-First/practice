from selenium import webdriver
from openpyxl import load_workbook, Workbook
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import StaleElementReferenceException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions
from openpyxl.styles import Font
from selenium import webdriver
from openpyxl import load_workbook, Workbook

def namuwiki_crawl(workbook, file_name, url, sheet):
    chromedriver = "C:\hanna\practice\for-hanna\chromedriver.exe"
    driver = webdriver.Chrome(chromedriver)
    # url = "https://namu.wiki/w/%EB%84%A4%EC%9D%B4%EB%B2%84%20%EC%9B%B9%ED%88%B0/%EC%97%B0%EC%9E%AC%20%EB%AA%A9%EB%A1%9D"
    driver.get(url)

    #웹툰 정보
    webtoon_genre = '//*[@id="wd8ZpUELU"]/article/div[4]/div/div/div/div/div/div[6]/div/div/div/div/div/div/div/div/div/div[6]/div[1]/div/div/table/tbody/tr[3]/td[2]/div/a'  
        webtoon_check_genre = '//*[@id="wd8ZpUELU"]/article/div[4]/div/div/div/div/div/div[6]/div/div/div/div/div/div/div/div/div/div[6]/div[1]/div/div/table/tbody/tr[3]/td[1]/div/strong/span'

# if webtoon_check_genre == "장르":
#     print("테스트 완") # 테스트 통과

    webtoon_genre = '//*[@id="wd8ZpUELU"]/article/div[4]/div/div/div/div/div/div[6]/div/div/div/div/div/div/div/div/div/div[6]/div[1]/div/div/table/tbody/tr[3]/td[2]/div'  
    text = driver.find_elements_by_xpath(webtoon_genre) #ts물, 로맨스, ..

    check_genre = ""
    total_genre = ""
    try:
        check_genre = driver.find_element_by_xpath(webtoon_check_genre).text
        if len(text) > 0:
            if check_genre == "장르":
                genre_list = []
                cnt = 0
                try:
                    for zz in text:
                        if cnt == 0:
                            cnt += 1
                            total_genre = zz.text
                        else:
                            genre_list.append(zz.text)
                            total_genre = list(total_genre) + genre_list
                except ValueError:
                    pass
    except NoSuchElementException:
        pass
    except StaleElementReferenceException:
        print("StaleElementReferenceException")
        return total_genre
    print("total_genre : ", total_genre)
    driver.quit()

    sheet['B' + str(cc)] = total_genre
    workbook.save(file_name)
    return total_genre

#기존 네이버 웹툰 url 엑셀 파일 불러오기
file_name = "webtoon_mon.xlsx"
workbook = load_workbook(file_name)
sheet = workbook['Sheet']
cc = 1
for i in sheet['B']:
    url = i.value
    namuwiki_crawl(workbook, file_name, url, sheet)
    cc += 1
